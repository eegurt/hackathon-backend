from rest_framework import viewsets, filters
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from io import BytesIO
from openpyxl import Workbook, load_workbook
from .models import ResourceType, WaterType, Object, PriorityScore, Region
from .serializer import (
    RegionSerializer,
    ResourceTypeSerializer,
    WaterTypeSerializer,
    ObjectSerializer,
    PriorityScoreSerializer,
)
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from .services.priority import calculate_priority_score
from .models import PriorityLevel
from datetime import date
import django_filters


class ObjectFilter(django_filters.FilterSet):
    passport_date = django_filters.DateFromToRangeFilter(field_name="passport_date")

    class Meta:
        model = Object
        fields = {
            "region": ["exact"],
            "resource_type": ["exact"],
            "water_type": ["exact"],
            "fauna": ["exact"],
            "technical_condition": ["exact"],
        }


class RegionViewSet(viewsets.ModelViewSet):
    queryset = Region.objects.all()
    serializer_class = RegionSerializer


class ResourceTypeViewSet(viewsets.ModelViewSet):
    queryset = ResourceType.objects.all()
    serializer_class = ResourceTypeSerializer


class WaterTypeViewSet(viewsets.ModelViewSet):
    queryset = WaterType.objects.all()
    serializer_class = WaterTypeSerializer


class ObjectViewSet(viewsets.ModelViewSet):
    queryset = Object.objects.all()
    serializer_class = ObjectSerializer
    filter_backends = [filters.SearchFilter, django_filters.rest_framework.DjangoFilterBackend]
    search_fields = ["name"]
    filterset_class = ObjectFilter

    @action(detail=False, methods=["get"])
    def export_xls(self, request):
        """
        Экспорт объектов в XLSX.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Objects"

        headers = [
            "id",
            "name",
            "region_id",
            "resource_type_id",
            "water_type_id",
            "fauna",
            "passport_date",
            "technical_condition",
            "latitude",
            "longitude",
            "pdf",
            "priority",
            "created_at",
        ]
        ws.append(headers)

        queryset = Object.objects.select_related("region", "resource_type", "water_type")
        for obj in queryset:
            ws.append([
                obj.id,
                obj.name,
                obj.region_id,
                obj.resource_type_id,
                obj.water_type_id,
                obj.fauna,
                obj.passport_date.isoformat() if obj.passport_date else None,
                obj.technical_condition,
                float(obj.latitude) if obj.latitude is not None else None,
                float(obj.longitude) if obj.longitude is not None else None,
                obj.pdf.name if obj.pdf else "",
                obj.priority,
                obj.created_at.isoformat() if obj.created_at else None,
            ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="objects.xlsx"'
        return response

    @action(detail=False, methods=["post"])
    def import_xls(self, request):
        """
        Импорт объектов из XLSX.
        Ожидается файл в `file` с колонками как у экспорта.
        """
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "file is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(upload)
            ws = wb.active
        except Exception as exc:
            return Response({"detail": f"Invalid XLSX file: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        created = 0
        updated = 0
        errors = []

        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
        expected_headers = [
            "id",
            "name",
            "region_id",
            "resource_type_id",
            "water_type_id",
            "fauna",
            "passport_date",
            "technical_condition",
            "latitude",
            "longitude",
            "pdf",
            "priority",
            "created_at",
        ]
        if header != expected_headers:
            return Response({"detail": "Unexpected headers", "expected": expected_headers, "got": header}, status=status.HTTP_400_BAD_REQUEST)

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # пропускаем пустые строки
            if all(cell is None for cell in row):
                continue
            try:
                (
                    obj_id,
                    name,
                    region_id,
                    resource_type_id,
                    water_type_id,
                    fauna,
                    passport_date,
                    technical_condition,
                    latitude,
                    longitude,
                    pdf_name,
                    priority,
                    _created_at,
                ) = row

                region = Region.objects.get(pk=region_id)
                resource_type = ResourceType.objects.get(pk=resource_type_id)
                water_type = WaterType.objects.get(pk=water_type_id)

                obj_data = {
                    "name": name,
                    "region": region,
                    "resource_type": resource_type,
                    "water_type": water_type,
                    "fauna": bool(fauna),
                    "passport_date": passport_date,
                    "technical_condition": int(technical_condition) if technical_condition is not None else 0,
                    "latitude": latitude,
                    "longitude": longitude,
                    "priority": int(priority) if priority is not None else 0,
                }

                if obj_id:
                    obj = Object.objects.get(pk=obj_id)
                    for field, value in obj_data.items():
                        setattr(obj, field, value)
                    obj.save()
                    updated += 1
                else:
                    Object.objects.create(**obj_data)
                    created += 1
            except Exception as exc:
                errors.append({"row": idx, "error": str(exc)})

        return Response({
            "created": created,
            "updated": updated,
            "errors": errors,
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    def recalc_priority(self, request, pk=None):
        obj = self.get_object()

        # считаем новую оценку
        score = calculate_priority_score(obj, today=date.today())

        # классификация
        if score >= 12:
            level = PriorityLevel.HIGH
        elif score >= 6:
            level = PriorityLevel.MEDIUM
        else:
            level = PriorityLevel.LOW

        priority, _ = PriorityScore.objects.get_or_create(obj=obj)
        priority.score = score
        priority.level = level
        priority.save()

        # синхронизируем старое поле приоритета, если оно ещё используется
        obj.priority = score
        obj.save(update_fields=["priority"])

        return Response({
            "object_id": obj.id,
            "new_score": score,
            "level": level,
            "status": "updated"
        }, status=status.HTTP_200_OK)


class PriorityScoreViewSet(viewsets.ModelViewSet):
    queryset = PriorityScore.objects.all()
    serializer_class = PriorityScoreSerializer

    @action(detail=False, methods=["post"])
    def recalc(self, request):
        """
        Пересчитать приоритет для объекта по его ID (для фронта).
        """
        object_id = request.data.get("object_id")
        if not object_id:
            return Response({"detail": "object_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(Object, pk=object_id)

        score = calculate_priority_score(obj, today=date.today())

        if score >= 12:
            level = PriorityLevel.HIGH
        elif score >= 6:
            level = PriorityLevel.MEDIUM
        else:
            level = PriorityLevel.LOW

        priority_obj, _ = PriorityScore.objects.get_or_create(obj=obj)
        priority_obj.score = score
        priority_obj.level = level
        priority_obj.save()

        obj.priority = score
        obj.save(update_fields=["priority"])

        return Response({
            "object_id": obj.id,
            "new_score": score,
            "level": level,
            "status": "updated"
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="by-object")
    def get_by_object(self, request, pk=None):
        """
        Получить PriorityScore по ID объекта (obj_id).
        """
        obj = get_object_or_404(Object, pk=pk)
        priority, _ = PriorityScore.objects.get_or_create(obj=obj)
        serializer = self.get_serializer(priority)
        return Response(serializer.data, status=status.HTTP_200_OK)
