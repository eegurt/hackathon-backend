from django.contrib import admin, messages
from django.template.response import TemplateResponse
from django.urls import path
from django.shortcuts import redirect
from django.http import HttpResponse, HttpResponseForbidden
from io import BytesIO
from datetime import date
from openpyxl import Workbook, load_workbook
from .models import ResourceType, WaterType, Object, PriorityScore, Region
from .services.priority import calculate_priority_score
from .models import PriorityLevel


@admin.action(description="Пересчитать приоритет для выбранных объектов")
def recalc_priority(modeladmin, request, queryset):
    for obj in queryset:
        score = calculate_priority_score(obj, today=date.today())

        if score >= 12:
            level = PriorityLevel.HIGH
        elif score >= 6:
            level = PriorityLevel.MEDIUM
        else:
            level = PriorityLevel.LOW

        priority_obj, _ = PriorityScore.objects.get_or_create(obj=obj)
        priority_obj.score = score
        priority_obj.level = level
        priority_obj.save()

        obj.priority = score
        obj.save(update_fields=["priority"])


@admin.action(description="Экспортировать выбранные объекты в XLSX")
def export_objects_xls(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Objects"

    headers = [
        "id",
        "name",
        "region_id",
        "resource_type_id",
        "water_type_id",
        "fauna",
        "passport_date",
        "technical_condition",
        "latitude",
        "longitude",
        "pdf",
        "priority",
        "created_at",
    ]
    ws.append(headers)

    for obj in queryset.select_related("region", "resource_type", "water_type"):
        ws.append([
            obj.id,
            obj.name,
            obj.region_id,
            obj.resource_type_id,
            obj.water_type_id,
            obj.fauna,
            obj.passport_date.isoformat() if obj.passport_date else None,
            obj.technical_condition,
            float(obj.latitude) if obj.latitude is not None else None,
            float(obj.longitude) if obj.longitude is not None else None,
            obj.pdf.name if obj.pdf else "",
            obj.priority,
            obj.created_at.isoformat() if obj.created_at else None,
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="objects.xlsx"'
    return response


@admin.register(Object)
class ObjectAdmin(admin.ModelAdmin):
    change_list_template = "admin/atla/object/change_list.html"
    actions = [recalc_priority, export_objects_xls]

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("import-xls/", self.admin_site.admin_view(self.import_xls_view), name="atla_object_import_xls"),
        ]
        return custom_urls + urls

    def import_xls_view(self, request):
        if not self.has_change_permission(request):
            return HttpResponseForbidden()

        if request.method == "POST":
            upload = request.FILES.get("file")
            if not upload:
                messages.error(request, "Нужно выбрать файл XLSX.")
                return redirect("..")

            try:
                wb = load_workbook(upload)
                ws = wb.active
            except Exception as exc:
                messages.error(request, f"Некорректный файл XLSX: {exc}")
                return redirect("..")

            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
            expected_headers = [
                "id",
                "name",
                "region_id",
                "resource_type_id",
                "water_type_id",
                "fauna",
                "passport_date",
                "technical_condition",
                "latitude",
                "longitude",
                "pdf",
                "priority",
                "created_at",
            ]
            if header != expected_headers:
                messages.error(request, f"Неверные заголовки. Ожидались: {expected_headers}. Получены: {header}")
                return redirect("..")

            created = 0
            updated = 0
            errors = []

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(cell is None for cell in row):
                    continue
                try:
                    (
                        obj_id,
                        name,
                        region_id,
                        resource_type_id,
                        water_type_id,
                        fauna,
                        passport_date,
                        technical_condition,
                        latitude,
                        longitude,
                        _pdf_name,
                        priority,
                        _created_at,
                    ) = row

                    region = Region.objects.get(pk=region_id)
                    resource_type = ResourceType.objects.get(pk=resource_type_id)
                    water_type = WaterType.objects.get(pk=water_type_id)

                    obj_data = {
                        "name": name,
                        "region": region,
                        "resource_type": resource_type,
                        "water_type": water_type,
                        "fauna": bool(fauna),
                        "passport_date": passport_date,
                        "technical_condition": int(technical_condition) if technical_condition is not None else 0,
                        "latitude": latitude,
                        "longitude": longitude,
                        "priority": int(priority) if priority is not None else 0,
                    }

                    if obj_id:
                        obj = Object.objects.get(pk=obj_id)
                        for field, value in obj_data.items():
                            setattr(obj, field, value)
                        obj.save()
                        updated += 1
                    else:
                        Object.objects.create(**obj_data)
                        created += 1
                except Exception as exc:
                    errors.append({"row": idx, "error": str(exc)})

            if created or updated:
                messages.success(request, f"Импорт завершён: создано {created}, обновлено {updated}.")
            if errors:
                messages.warning(request, f"Ошибки в {len(errors)} строках. Пример: {errors[0]}")

            return redirect("..")

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Импорт объектов из XLSX",
        }
        return TemplateResponse(request, "admin/atla/object/import_xls.html", context)


admin.site.register(ResourceType)
admin.site.register(WaterType)
admin.site.register(PriorityScore)
